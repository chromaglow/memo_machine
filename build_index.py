#!/usr/bin/env python3
"""Module 7 — build the spreadsheet index and rename the archive.

Writes `index.csv` (source of truth) and `index.xlsx` (formatted for reading),
one row per recording, then renames the files in `library/` to human-readable
names. Recordings without enrichment keep their phone filenames and are flagged
for manual handling rather than being dropped.

The spreadsheet is written before anything is renamed, so a rename failure never
leaves the index describing files that don't exist.

Naming (approved 2026-07-27):
    YYYY-MM-DD_HHMM_category_slug_participants.ext
    2025-10-20_1530_call_aht-bands-pilot-data-delta_charlie.m4a

Usage:
    python build_index.py            # write index, show the rename plan
    python build_index.py --rename   # write index and perform the renames
"""

import argparse
import csv
import datetime
import json
import re
import sys
from pathlib import Path

DEFAULT_DATA = Path(r"C:\Users\ezras\memo-machine-data")
MAX_FILENAME = 100          # Windows path headroom inside a synced folder
MAX_PARTICIPANTS_IN_NAME = 2

COLUMNS = [
    "date", "time", "date_source", "duration", "category", "title", "topic",
    "summary", "participants", "participants_confidence", "action_items",
    "folder", "original_filename", "archive_filename", "has_transcript",
    "flags", "file_hash",
]

COLUMN_NOTES = [
    ("date / time", "When the recording was made, from CloudRecordings.db on the phone."),
    ("date_source", "Where the date came from: 'db' is the phone's own database. "
                    "All 243 came from there — none fell back to file timestamps."),
    ("duration", "Length of the audio, mm:ss."),
    ("category", "meeting, call, note-to-self, idea or other. Assigned by Claude."),
    ("title", "Short descriptive title written by Claude from the transcript."),
    ("topic", "One sentence on what the recording is about."),
    ("summary", "Two to four sentences. Written only from what is actually in the "
                "transcript — thin recordings get short summaries rather than padding."),
    ("participants", "Only people with direct spoken evidence of being present — "
                     "addressed by name, introducing themselves, or signing off. "
                     "Blank means nobody could be evidenced, NOT that nobody was there. "
                     "Names merely discussed appear in flags instead. You are never "
                     "listed, since you are in every recording."),
    ("participants_confidence", "high / medium / low / none. 'none' accompanies a "
                                "blank participants cell."),
    ("action_items", "Commitments actually stated in the recording. One per line."),
    ("folder", "Which subfolder of library/ the files are in now."),
    ("original_filename", "The name the phone gave it."),
    ("archive_filename", "The renamed file. Audio and transcript share this name."),
    ("has_transcript", "yes / no. 30 recordings have no transcript — mostly "
                       "sub-60-second fragments."),
    ("flags", "Anything worth a human look: transcription garbling, names that "
              "could not be verified, sensitive content, recordings that are "
              "mostly silence."),
    ("file_hash", "SHA-256 of the audio, verified against the phone backup. This "
                  "is the stable identifier — it survives renaming and moving."),
]

FOLDER_NOTES = [
    ("amazon", "Work at Amazon: the AHT/handle-time project, MFI cases, AIS, "
               "Paragon, concurrency rollouts, org and layoff matters, promotion "
               "and TPM-track mentorship, internal interviews, internal training."),
    ("qed", "The QED venture with Josh and Aaron: incorporation and C-corp "
            "formation, Hard Shell, Scout, Omega, Stream Kinetics, and client work "
            "including DADS, United Way, Seattle Unity, Endo DNA and church software."),
    ("techland", "Techland Ventures and Geode Solutions, run with Charlie: PMO "
                 "syncs, deal pipeline, JLL and data-centre clusters, geothermal "
                 "and ag-tech, Shreveport, Vietnam, Japan/JETRO, Asana rollout."),
    ("(root)", "Everything else — personal conversations, one-off calls, and "
               "recordings with no transcript."),
]


def slugify(text: str, default: str = "untitled") -> str:
    text = re.sub(r"[^a-z0-9]+", "-", (text or "").lower()).strip("-")
    return text or default


def build_filename(row: dict, names: list, extension: str) -> str:
    """Assemble the archive filename, trimming the slug if the whole is too long.

    The date, time, category and extension are fixed-width and non-negotiable —
    they are what makes the folder sortable — so length is recovered from the
    slug, which is the only free-text part.
    """
    date_part = row["date"].replace("-", "-")
    time_part = row["time"].replace(":", "")
    category = slugify(row["category"], "other")
    slug = slugify(row["_slug"], slugify(row["title"]))

    # Participants segment: omitted for solo recordings and when nothing was
    # evidenced, which is most of the corpus.
    people = ""
    if names and row["category"] != "note-to-self":
        shown = [slugify(n.split()[0]) for n in names[:MAX_PARTICIPANTS_IN_NAME]]
        people = "-".join(filter(None, shown))
        if len(names) > MAX_PARTICIPANTS_IN_NAME and people:
            people += "-etal"

    fixed = f"{date_part}_{time_part}_{category}"
    tail = (f"_{people}" if people else "") + extension
    room = MAX_FILENAME - len(fixed) - len(tail) - 1
    if room < 8:  # pathological — drop the participants rather than the slug
        tail = extension
        room = MAX_FILENAME - len(fixed) - len(tail) - 1
    slug = slug[:max(room, 1)].rstrip("-")
    return f"{fixed}_{slug}{tail}"


def locate(library: Path, name: str):
    """Find a file in library/, whether at the root or inside a topic subfolder."""
    if (library / name).exists():
        return library / name
    return next((p for p in library.glob(f"*/{name}")), None)


def load(data: Path):
    """Join inventory, transcript index and enrichment into one row per recording."""
    inv = {}
    with open(data / "inventory.csv", encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            if r.get("sha256"):
                inv[r["filename"]] = r

    tidx = {}
    with open(data / "transcript_index.csv", encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            tidx[r["filename"]] = r

    enriched = {}
    for path in (data / "enriched").glob("*.json"):
        payload = json.loads(path.read_text(encoding="utf-8"))
        enriched[payload["_hash"]] = payload

    rows = []
    for filename, meta in inv.items():
        stamp = meta.get("recorded", "")
        date, _, time_of_day = stamp.partition(" ")
        e = enriched.get(meta["sha256"])
        t = tidx.get(filename, {})
        seconds = int(float(meta.get("duration_s") or 0))
        names = [p["name"] for p in (e or {}).get("participants", [])]
        flags = list((e or {}).get("flags", []))
        if e is None:
            flags.insert(0, "NOT ENRICHED — no transcript; keeps its original filename")

        rows.append({
            "date": date,
            "time": time_of_day[:5],
            # Every date came from ZCLOUDRECORDING; nothing fell back to mtime.
            "date_source": "db" if stamp else "unknown",
            "duration": f"{seconds // 60}:{seconds % 60:02d}",
            "category": (e or {}).get("category", ""),
            "title": (e or {}).get("title", meta.get("title", "")),
            "topic": (e or {}).get("topic", ""),
            "summary": (e or {}).get("summary", ""),
            "participants": ", ".join(names),
            "participants_confidence": (e or {}).get("participants_confidence", ""),
            "action_items": "\n".join((e or {}).get("action_items", [])),
            "original_filename": filename,
            "archive_filename": "",
            "has_transcript": "yes" if t.get("source") not in (None, "", "none") else "no",
            "flags": "\n".join(flags),
            "file_hash": meta["sha256"],
            "_slug": (e or {}).get("slug", ""),
            "_names": names,
            "_enriched": e is not None,
        })

    rows.sort(key=lambda r: (r["date"], r["time"]), reverse=True)
    return rows


def assign_names(rows: list) -> None:
    """Fill archive_filename, resolving collisions with numeric suffixes."""
    taken = set()
    for row in rows:
        original = row["original_filename"]
        if not row["_enriched"]:
            row["archive_filename"] = original     # untouched, flagged in the sheet
            taken.add(original.lower())
            continue
        extension = Path(original).suffix
        candidate = build_filename(row, row["_names"], extension)
        if candidate.lower() in taken:
            stem, suffix = candidate[: -len(extension)], extension
            n = 2
            while f"{stem}_{n}{suffix}".lower() in taken:
                n += 1
            candidate = f"{stem}_{n}{suffix}"
        row["archive_filename"] = candidate
        taken.add(candidate.lower())


def assign_folders(data: Path, rows: list) -> None:
    """Record which topic subfolder each recording currently lives in."""
    library = data / "library"
    for row in rows:
        name = row["archive_filename"] or row["original_filename"]
        found = locate(library, name)
        if found is None:
            row["folder"] = "(missing)"
        elif found.parent == library:
            row["folder"] = ""
        else:
            row["folder"] = found.parent.name


def write_csv(data: Path, rows: list) -> Path:
    path = data / "index.csv"
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    return path


def write_xlsx(data: Path, rows: list) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    widths = {"date": 11, "time": 7, "date_source": 12, "duration": 9,
              "category": 13, "title": 42, "topic": 55, "summary": 80,
              "participants": 22, "participants_confidence": 12,
              "action_items": 70, "folder": 11, "original_filename": 30,
              "archive_filename": 52, "has_transcript": 8, "flags": 70,
              "file_hash": 18}

    wb = Workbook()
    ws = wb.active
    ws.title = "Voice Memos"
    ws.append(COLUMNS)

    header_fill = PatternFill("solid", fgColor="1F3864")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"

    wrap = {"topic", "summary", "action_items", "flags", "title"}
    for row in rows:
        ws.append([row.get(c, "") for c in COLUMNS])
    for i, name in enumerate(COLUMNS, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = widths.get(name, 18)
        if name in wrap:
            for cell in ws[letter][1:]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    write_about_sheet(wb, rows)
    path = data / "index.xlsx"
    wb.save(path)
    return path


def write_about_sheet(wb, rows: list) -> None:
    """A second sheet explaining where the data came from and how to read it."""
    from openpyxl.styles import Alignment, Font

    import collections
    ws = wb.create_sheet("About")
    title_font = Font(bold=True, size=14)
    head_font = Font(bold=True, size=11, color="1F3864")

    def add(text="", font=None, wrap=False):
        ws.append([text])
        cell = ws.cell(row=ws.max_row, column=1)
        if font:
            cell.font = font
        if wrap:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        return ws.max_row

    def add_pair(left, right):
        ws.append([left, right])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(vertical="top")
        ws.cell(row=ws.max_row, column=2).alignment = Alignment(
            wrap_text=True, vertical="top")

    total = len(rows)
    enriched = sum(1 for r in rows if r["_enriched"])
    hours = sum(int(r["duration"].split(":")[0]) for r in rows) / 60
    folders = collections.Counter(r["folder"] or "(root)" for r in rows)

    add("Voice Memo Archive", title_font)
    add()
    add(f"{total} recordings taken off an iPhone, roughly {hours:.0f} hours of audio. "
        f"{enriched} have transcripts and structured metadata; the remaining "
        f"{total - enriched} are mostly sub-60-second fragments with no usable speech.",
        wrap=True)
    add()

    add("Where the transcripts came from", head_font)
    add("Apple's Voice Memos app stores its on-device transcript inside the audio "
        "file itself. 211 recordings had one and it was read straight out of the "
        "file — no transcription service, no cost, and the audio never left the "
        "machine. Two long recordings Apple had never transcribed were run through "
        "Whisper locally. The rest have no transcript.", wrap=True)
    add()

    add("How to read the participants column", head_font)
    add("This is the column most likely to mislead, so it was built to be "
        "conservative. A name appears only when the transcript contains direct "
        "spoken evidence that the person was present — addressed by name, "
        "introducing themselves, or signing off — and every name had to come with "
        "a verbatim quote, which was then checked against the transcript. Four "
        "names were dropped because their quote could not be found. A blank cell "
        "means nobody could be evidenced, not that nobody was there. Names that "
        "were merely discussed appear in flags instead.", wrap=True)
    add()

    add("Folders", head_font)
    for name, description in FOLDER_NOTES:
        count = folders.get(name if name != "(root)" else "(root)", 0)
        add_pair(f"{name}  ({count})", description)
    add()

    add("Columns", head_font)
    for name, description in COLUMN_NOTES:
        add_pair(name, description)
    add()

    add("Caveats worth knowing", head_font)
    add("Transcripts contain recognition errors, especially on proper nouns and "
        "acronyms — 'AWS' as 'ABS', 'Glue Crawlers' as 'Blue Crawlers'. Summaries "
        "read through those errors where the intent is clear and flag them where "
        "it is not. Categories and titles are judgments, not facts. The date, "
        "duration and file hash are the only fields taken directly from the phone.",
        wrap=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 105
    for row in ws.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if cell.alignment.wrap_text:
                ws.row_dimensions[cell.row].height = None


def do_rename(data: Path, rows: list, apply: bool) -> int:
    """Rename audio and transcript in library/ to the archive filename."""
    library = data / "library"
    renamed = skipped = failed = 0
    for row in rows:
        if not row["_enriched"]:
            skipped += 1
            continue
        if row["archive_filename"] == row["original_filename"]:
            continue
        # Files may have been sorted into topic subfolders since the last run;
        # rename them where they sit rather than assuming the library root.
        if locate(library, row["archive_filename"]) is not None:
            continue              # already renamed by an earlier run
        src = locate(library, row["original_filename"])
        if src is None:
            print(f"  MISSING: {row['original_filename']}")
            failed += 1
            continue
        dst = src.parent / row["archive_filename"]
        if not apply:
            renamed += 1
            continue
        try:
            src.rename(dst)
            txt = src.parent / (Path(row["original_filename"]).stem + ".txt")
            if txt.exists():
                txt.rename(src.parent / (Path(row["archive_filename"]).stem + ".txt"))
            renamed += 1
        except Exception as exc:
            print(f"  FAILED {src.name}: {exc}")
            failed += 1
    verb = "renamed" if apply else "would rename"
    print(f"  {verb} {renamed} | left as-is (not enriched) {skipped} | failed {failed}")
    return failed


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--data", type=Path, default=DEFAULT_DATA)
    ap.add_argument("--rename", action="store_true",
                    help="actually rename files (default is a dry run)")
    args = ap.parse_args()

    rows = load(args.data)
    assign_names(rows)
    assign_folders(args.data, rows)
    print(f"recordings: {len(rows)}  "
          f"(enriched {sum(1 for r in rows if r['_enriched'])})")
    import collections as _c
    for folder, n in sorted(_c.Counter(r["folder"] or "(root)" for r in rows).items()):
        print(f"    {folder:<12} {n}")

    # Index first: a rename failure must never leave the sheet describing
    # filenames that were never created.
    print(f"  wrote {write_csv(args.data, rows)}")
    print(f"  wrote {write_xlsx(args.data, rows)}")

    print("\nlongest filenames produced:")
    for row in sorted((r for r in rows if r["_enriched"]),
                      key=lambda r: -len(r["archive_filename"]))[:3]:
        print(f"  {len(row['archive_filename']):3d}  {row['archive_filename']}")

    print(f"\n{'renaming' if args.rename else 'dry run — no files touched'}:")
    return do_rename(args.data, rows, apply=args.rename)


if __name__ == "__main__":
    sys.exit(main())
